"""
Chunking Intelligent avec Fallbacks Multiples
Adapté pour LLM local (Mistral via Ollama)
"""

import sys
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import logging
from bs4 import BeautifulSoup
import re

# Ajouter chemin utils
_project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(_project_root / 'src' / 'utils'))
from rgpd_topics import parse_tags, TAG_PROMPT_TABLE

logger = logging.getLogger(__name__)


class ChunkingConfig:
    """Configuration chunking adaptée Mistral local"""
    
    # Limites LLM
    MAX_CONTEXT_WINDOW = 32768      # tokens Mistral
    SAFE_MARGIN = 0.7               # Utiliser 70% pour sécurité
    RESERVE_OUTPUT = 2000           # Tokens pour réponse LLM
    RESERVE_PROMPT = 500            # Tokens pour prompt
    TOKENS_PER_WORD_FR = 1.3        # Approximation français
    
    # Calculs automatiques
    MAX_INPUT_TOKENS = int(MAX_CONTEXT_WINDOW * SAFE_MARGIN)
    MAX_CHUNK_TOKENS = MAX_INPUT_TOKENS - RESERVE_OUTPUT - RESERVE_PROMPT
    HARD_LIMIT_WORDS = int(MAX_CHUNK_TOKENS / TOKENS_PER_WORD_FR)  # ~15700 mots
    
    # Tailles cibles (mots)
    TARGET_CHUNK_SIZE = 500         # Idéal
    MIN_CHUNK_SIZE = 200            # Minimum viable
    MAX_CHUNK_SIZE = 1000           # Maximum avant split
    OVERLAP = 50                    # Overlap entre chunks
    
    # Stratégies par type
    STRATEGIES = {
        'html': 'semantic_hierarchy',
        'pdf': 'structure_detect',
        'ods': 'by_sheet',
        'xlsx': 'by_sheet',
        'docx': 'by_section',
        'odt': 'by_section',
    }


class ChunkMetadata:
    """Metadata enrichie pour un chunk"""
    
    def __init__(self, 
                 chunk_id: str,
                 parent_doc: str,
                 chunk_text: str,
                 chunk_type: str = 'semantic',
                 **kwargs):
        self.chunk_id = chunk_id
        self.parent_doc = parent_doc
        self.chunk_text = chunk_text
        self.chunk_type = chunk_type
        self.word_count = len(chunk_text.split())
        
        # Métadonnées optionnelles
        self.section = kwargs.get('section', '')
        self.section_summary = kwargs.get('section_summary', '')
        self.document_summary = kwargs.get('document_summary', '')
        self.position = kwargs.get('position', '')  # "1/10"
        self.is_continuation = kwargs.get('is_continuation', False)
        self.related_chunks = kwargs.get('related_chunks', [])
        
    def to_dict(self) -> Dict:
        return {
            'chunk_id': self.chunk_id,
            'parent_doc': self.parent_doc,
            'chunk_text': self.chunk_text,
            'chunk_type': self.chunk_type,
            'word_count': self.word_count,
            'section': self.section,
            'section_summary': self.section_summary,
            'document_summary': self.document_summary,
            'position': self.position,
            'is_continuation': self.is_continuation,
            'related_chunks': self.related_chunks,
        }


class SemanticChunker:
    """Chunking sémantique intelligent avec fallbacks"""
    
    def __init__(self, config: Optional[ChunkingConfig] = None):
        self.config = config or ChunkingConfig()
    
    def chunk_document(self, 
                      file_path: Path, 
                      doc_metadata: Dict) -> List[ChunkMetadata]:
        """Point d'entrée : chunke un document selon son type"""
        
        file_type = file_path.suffix.lower()[1:]  # Sans le point
        
        strategy = self.config.STRATEGIES.get(file_type, 'fallback')
        
        logger.info(f"📄 Chunking {file_path.name} avec stratégie: {strategy}")
        
        if file_type == 'html':
            return self.chunk_html(file_path, doc_metadata)
        elif file_type == 'pdf':
            return self.chunk_pdf(file_path, doc_metadata)
        elif file_type in ['ods', 'xlsx']:
            return self.chunk_spreadsheet(file_path, doc_metadata)
        elif file_type in ['docx', 'odt']:
            return self.chunk_document_text(file_path, doc_metadata)
        else:
            logger.warning(f"⚠️  Type inconnu {file_type}, fallback split fixe")
            return self.chunk_fallback(file_path, doc_metadata)
    
    # =========================================================================
    # HTML : Chunking Hiérarchique Sémantique
    # =========================================================================
    
    def chunk_html(self, file_path: Path, doc_metadata: Dict) -> List[ChunkMetadata]:
        """Chunk HTML par structure sémantique — cible le contenu principal"""
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            soup = BeautifulSoup(f, 'lxml')
        
        # Cibler le bloc de contenu principal (structure CNIL)
        content_root = (
            soup.find(class_='region-content')
            or soup.find('main')
            or soup.find('article')
            or soup.find(class_='field-name-body')
            or soup  # Fallback
        )
        
        # Supprimer bruit (navigation, menus, etc.)
        for tag in content_root(['script', 'style', 'nav', 'header', 'footer', 
                                  'aside', 'iframe', 'noscript', 'svg']):
            tag.decompose()
        
        # Supprimer menus, breadcrumbs, pagination
        for nav_block in content_root.find_all(class_=lambda c: c and any(
            x in str(c).lower() for x in [
                'menu-push', 'breadcrumb', 'pager', 'pagination',
                'nav-', 'share-', 'social', 'cookie', 'back-to-top'
            ]
        )):
            nav_block.decompose()
        
        chunks = []
        
        # Stratégie 1: Par <article> dans le contenu
        articles = content_root.find_all('article')
        if articles:
            logger.debug(f"  ✅ {len(articles)} <article> trouvés")
            for i, article in enumerate(articles):
                chunks.extend(self._chunk_html_element(
                    article, 
                    doc_metadata, 
                    prefix=f"article_{i}",
                    level="article"
                ))
            return chunks
        
        # Stratégie 2: Par <section>
        sections = content_root.find_all('section')
        if sections:
            logger.debug(f"  ✅ {len(sections)} <section> trouvés")
            for i, section in enumerate(sections):
                chunks.extend(self._chunk_html_element(
                    section,
                    doc_metadata,
                    prefix=f"section_{i}",
                    level="section"
                ))
            return chunks
        
        # Stratégie 3: Par <div> avec classe content/main
        main_divs = content_root.find_all('div', class_=re.compile(r'(content|main|article|post)'))
        if main_divs:
            logger.debug(f"  ✅ {len(main_divs)} <div> de contenu trouvés")
            for i, div in enumerate(main_divs):
                chunks.extend(self._chunk_html_element(
                    div,
                    doc_metadata,
                    prefix=f"content_{i}",
                    level="div"
                ))
            return chunks
        
        # Fallback: Tout le contenu ciblé
        logger.warning(f"  ⚠️  Pas de structure claire, chunking du contenu principal")
        text = content_root.get_text(separator=' ', strip=True)
        return self._split_text_fixed(text, doc_metadata, "body")
    
    def _chunk_html_element(self, 
                           element, 
                           doc_metadata: Dict,
                           prefix: str,
                           level: str) -> List[ChunkMetadata]:
        """Chunk un élément HTML (article, section, div)"""
        
        text = element.get_text(separator=' ', strip=True)
        word_count = len(text.split())
        
        # Extraire titre si présent
        title_tag = element.find(['h1', 'h2', 'h3'])
        section_title = title_tag.get_text(strip=True) if title_tag else ''
        
        # Cas 1: Taille OK → 1 chunk
        if word_count <= self.config.MAX_CHUNK_SIZE:
            return [ChunkMetadata(
                chunk_id=f"{doc_metadata['doc_id']}_{prefix}",
                parent_doc=doc_metadata['doc_id'],
                chunk_text=text,
                chunk_type=f"semantic_{level}",
                section=section_title,
                position="1/1",
            )]
        
        # Cas 2: Trop gros → Split par sous-titres
        logger.debug(f"  ⚠️  {level} trop gros ({word_count} mots), split par sous-titres")
        
        subsections = element.find_all(['h2', 'h3', 'h4'])
        if subsections:
            return self._chunk_by_headings(element, subsections, doc_metadata, prefix, section_title)
        
        # Cas 3: Pas de sous-titres → Split fixe
        logger.debug(f"  ⚠️  Pas de sous-titres, split fixe")
        return self._split_text_fixed(text, doc_metadata, prefix, section_title)
    
    def _chunk_by_headings(self, 
                          element,
                          headings: List,
                          doc_metadata: Dict,
                          prefix: str,
                          parent_section: str) -> List[ChunkMetadata]:
        """Split HTML par sous-titres (h2, h3, h4) avec tracking relations"""
        
        chunks = []
        current_text = []
        current_heading = parent_section
        
        for child in element.children:
            if child.name in ['h2', 'h3', 'h4']:
                # Sauver chunk précédent
                if current_text:
                    text = ' '.join(current_text)
                    if len(text.split()) >= self.config.MIN_CHUNK_SIZE:
                        chunk_id = f"{doc_metadata['doc_id']}_{prefix}_h{len(chunks)}"
                        
                        # Related chunks (chunk précédent)
                        related = []
                        if len(chunks) > 0:
                            related.append(chunks[-1].chunk_id)
                        
                        chunks.append(ChunkMetadata(
                            chunk_id=chunk_id,
                            parent_doc=doc_metadata['doc_id'],
                            chunk_text=text,
                            chunk_type="semantic_subsection",
                            section=current_heading,
                            position=f"{len(chunks)+1}/?",
                            related_chunks=related,
                        ))
                
                # Nouveau chunk
                current_heading = child.get_text(strip=True)
                current_text = [current_heading]
            
            elif hasattr(child, 'get_text'):
                current_text.append(child.get_text(separator=' ', strip=True))
        
        # Dernier chunk
        if current_text:
            text = ' '.join(current_text)
            if len(text.split()) >= self.config.MIN_CHUNK_SIZE:
                chunk_id = f"{doc_metadata['doc_id']}_{prefix}_h{len(chunks)}"
                
                related = []
                if len(chunks) > 0:
                    related.append(chunks[-1].chunk_id)
                
                chunks.append(ChunkMetadata(
                    chunk_id=chunk_id,
                    parent_doc=doc_metadata['doc_id'],
                    chunk_text=text,
                    chunk_type="semantic_subsection",
                    section=current_heading,
                    position=f"{len(chunks)+1}/{len(chunks)+1}",
                    related_chunks=related,
                ))
        
        # Mettre à jour positions ET ajouter chunk suivant
        total = len(chunks)
        for i, chunk in enumerate(chunks):
            chunk.position = f"{i+1}/{total}"
            
            # Ajouter chunk suivant
            if i < total - 1:
                chunk.related_chunks.append(chunks[i+1].chunk_id)
        
        return chunks
    
    # =========================================================================
    # PDF : Détection Structure ou Fallback
    # =========================================================================
    
    def chunk_pdf(self, file_path: Path, doc_metadata: Dict) -> List[ChunkMetadata]:
        """Chunk PDF avec détection de structure et extraction robuste"""
        
        # Essayer extraction avec PyMuPDF (méthode préférée)
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(file_path)
            
            # Détecter structure (titres via TOC ou font size)
            toc = doc.get_toc()  # Table of contents
            
            if toc and len(toc) > 2:
                logger.debug(f"  ✅ Structure TOC détectée ({len(toc)} entrées)")
                chunks = self._chunk_pdf_by_toc(doc, toc, doc_metadata)
            else:
                logger.debug(f"  ⚠️  Pas de TOC, détection par font size")
                chunks = self._chunk_pdf_by_font(doc, doc_metadata)
            
            doc.close()
            return chunks
        
        except ImportError:
            logger.warning("⚠️  PyMuPDF non installé, essai pdfplumber")
        except Exception as e:
            logger.warning(f"⚠️  PyMuPDF échec ({e}), essai pdfplumber")
        
        # Fallback 1: pdfplumber
        try:
            import pdfplumber
            text_pages = []
            
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_pages.append(page_text)
            
            if text_pages:
                logger.debug(f"  ✅ Extraction pdfplumber: {len(text_pages)} pages")
                return self._chunk_text_pages(text_pages, doc_metadata)
        
        except ImportError:
            logger.warning("⚠️  pdfplumber non installé, essai PyPDF2")
        except Exception as e:
            logger.warning(f"⚠️  pdfplumber échec ({e}), essai PyPDF2")
        
        # Fallback 2: PyPDF2
        try:
            import PyPDF2
            text_pages = []
            
            with open(file_path, 'rb') as f:
                pdf = PyPDF2.PdfReader(f)
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_pages.append(page_text)
            
            if text_pages:
                logger.debug(f"  ✅ Extraction PyPDF2: {len(text_pages)} pages")
                return self._chunk_text_pages(text_pages, doc_metadata)
        
        except Exception as e:
            logger.warning(f"⚠️  PyPDF2 échec ({e}), essai OCR Tesseract")
        
        # Fallback 3: OCR Tesseract (PDFs scannés/images)
        try:
            import fitz
            from PIL import Image
            import pytesseract
            import io
            
            logger.info(f"  🔍 OCR Tesseract sur {file_path.name}...")
            
            doc = fitz.open(file_path)
            text_pages = []
            
            # Limiter à 10 premières pages pour OCR (très lent)
            max_pages_ocr = min(10, len(doc))
            
            for page_num in range(max_pages_ocr):
                page = doc[page_num]
                
                # Convertir page en image
                pix = page.get_pixmap(dpi=150)
                img_data = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_data))
                
                # OCR
                page_text = pytesseract.image_to_string(img, lang='fra')
                
                if page_text and len(page_text.strip()) > 50:
                    text_pages.append(page_text)
            
            doc.close()
            
            if text_pages:
                logger.debug(f"  ✅ OCR Tesseract: {len(text_pages)} pages extraites")
                return self._chunk_text_pages(text_pages, doc_metadata)
        
        except ImportError:
            logger.warning("⚠️  Tesseract non installé (pip install pytesseract)")
        except Exception as e:
            logger.error(f"❌ OCR Tesseract échec ({e})")
        
        # Fallback ultime: Créer chunk vide avec metadata
        logger.error(f"❌ TOUTES méthodes PDF échouées pour {file_path.name}")
        return [ChunkMetadata(
            chunk_id=f"{doc_metadata['doc_id']}_failed",
            parent_doc=doc_metadata['doc_id'],
            chunk_text=f"[PDF {file_path.name} - Extraction impossible avec toutes les méthodes]",
            chunk_type="failed_extraction",
            section="Extraction échouée",
            position="1/1",
        )]
    
    def _chunk_pdf_by_toc(self, 
                         doc, 
                         toc: List,
                         doc_metadata: Dict) -> List[ChunkMetadata]:
        """Chunk PDF en utilisant la table des matières"""
        
        chunks = []
        
        for i, (level, title, page_num) in enumerate(toc):
            # Extraire texte de cette section
            start_page = page_num - 1  # PyMuPDF indexe à 0
            
            # Trouver page de fin (prochaine section ou fin doc)
            if i < len(toc) - 1:
                end_page = toc[i + 1][2] - 1
            else:
                end_page = len(doc)
            
            # Extraire texte
            section_text = []
            for page_idx in range(start_page, min(end_page, len(doc))):
                page = doc[page_idx]
                section_text.append(page.get_text())
            
            text = '\n'.join(section_text)
            word_count = len(text.split())
            
            # Vérifier taille
            if word_count <= self.config.MAX_CHUNK_SIZE:
                chunks.append(ChunkMetadata(
                    chunk_id=f"{doc_metadata['doc_id']}_toc_{i}",
                    parent_doc=doc_metadata['doc_id'],
                    chunk_text=text,
                    chunk_type="semantic_toc",
                    section=title,
                    position=f"{i+1}/{len(toc)}",
                ))
            else:
                # Section trop grosse → split
                logger.debug(f"  ⚠️  Section '{title}' trop grosse ({word_count} mots), split")
                sub_chunks = self._split_text_fixed(text, doc_metadata, f"toc_{i}", title)
                chunks.extend(sub_chunks)
        
        return chunks
    
    def _chunk_pdf_by_font(self, 
                          doc,
                          doc_metadata: Dict) -> List[ChunkMetadata]:
        """Chunk PDF en détectant les titres par taille de police"""
        
        # Extraire tout le texte avec metadata de font
        blocks = []
        for page_num, page in enumerate(doc):
            blocks.extend(page.get_text("dict")["blocks"])
        
        # Détecter tailles de font majoritaires
        font_sizes = []
        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        font_sizes.append(span["size"])
        
        if not font_sizes:
            # Fallback: split par pages
            logger.warning(f"  ⚠️  Impossible de détecter font sizes, fallback pages")
            return self._chunk_pdf_by_pages(doc, doc_metadata)
        
        # Font médiane = texte normal
        median_size = sorted(font_sizes)[len(font_sizes) // 2]
        title_threshold = median_size * 1.2  # 20% plus grand = titre
        
        # Identifier sections
        sections = []
        current_section = {'title': '', 'text': []}
        
        for block in blocks:
            if "lines" not in block:
                continue
            
            for line in block["lines"]:
                line_text = ' '.join([span["text"] for span in line["spans"]])
                max_size = max([span["size"] for span in line["spans"]] or [0])
                
                if max_size >= title_threshold and len(line_text) < 200:
                    # C'est un titre
                    if current_section['text']:
                        sections.append(current_section)
                    current_section = {'title': line_text, 'text': []}
                else:
                    current_section['text'].append(line_text)
        
        # Dernier section
        if current_section['text']:
            sections.append(current_section)
        
        if not sections or len(sections) < 2:
            # Pas assez de structure → fallback pages
            logger.warning(f"  ⚠️  Structure insuffisante ({len(sections)} sections), fallback pages")
            return self._chunk_pdf_by_pages(doc, doc_metadata)
        
        # Créer chunks par section
        chunks = []
        for i, section in enumerate(sections):
            text = '\n'.join(section['text'])
            word_count = len(text.split())
            
            if word_count <= self.config.MAX_CHUNK_SIZE:
                chunks.append(ChunkMetadata(
                    chunk_id=f"{doc_metadata['doc_id']}_font_{i}",
                    parent_doc=doc_metadata['doc_id'],
                    chunk_text=text,
                    chunk_type="semantic_font",
                    section=section['title'],
                    position=f"{i+1}/{len(sections)}",
                ))
            else:
                sub_chunks = self._split_text_fixed(text, doc_metadata, f"font_{i}", section['title'])
                chunks.extend(sub_chunks)
        
        return chunks
    
    def _chunk_pdf_by_pages(self, 
                           doc,
                           doc_metadata: Dict) -> List[ChunkMetadata]:
        """Fallback: chunk PDF par groupes de pages"""
        
        # Calculer combien de pages par chunk
        total_text = []
        for page in doc:
            total_text.append(page.get_text())
        
        return self._chunk_text_pages(total_text, doc_metadata)
    
    def _chunk_text_pages(self, 
                         text_pages: List[str],
                         doc_metadata: Dict) -> List[ChunkMetadata]:
        """Chunk une liste de pages de texte (fallback pour PDF)"""
        
        total_words = sum(len(t.split()) for t in text_pages)
        avg_words_per_page = total_words / len(text_pages) if len(text_pages) > 0 else 300
        
        pages_per_chunk = max(1, int(self.config.TARGET_CHUNK_SIZE / avg_words_per_page))
        
        logger.debug(f"  📄 {len(text_pages)} pages, ~{avg_words_per_page:.0f} mots/page → {pages_per_chunk} pages/chunk")
        
        chunks = []
        overlap_pages = 1 if pages_per_chunk > 2 else 0  # 1 page overlap si groupes de 3+ pages
        
        i = 0
        while i < len(text_pages):
            # Extraire groupe de pages
            page_group = text_pages[i:i+pages_per_chunk]
            
            # Ajouter overlap de la page précédente si possible
            if i > 0 and overlap_pages > 0:
                overlap_text = text_pages[i-1]
                # Garder seulement les derniers N mots de l'overlap
                overlap_words = overlap_text.split()[-self.config.OVERLAP:]
                page_group.insert(0, ' '.join(overlap_words))
            
            text = '\n'.join(page_group)
            
            chunks.append(ChunkMetadata(
                chunk_id=f"{doc_metadata['doc_id']}_pages_{i//pages_per_chunk}",
                parent_doc=doc_metadata['doc_id'],
                chunk_text=text,
                chunk_type="fallback_pages",
                section=f"Pages {i+1}-{min(i+pages_per_chunk, len(text_pages))}",
                position=f"{len(chunks)+1}/?",
            ))
            
            # Avancer (sans re-compter l'overlap)
            i += pages_per_chunk
        
        # Update positions
        total = len(chunks)
        for i, chunk in enumerate(chunks):
            chunk.position = f"{i+1}/{total}"
            # Marquer les relations
            if i > 0:
                chunk.related_chunks.append(chunks[i-1].chunk_id)
            if i < total - 1:
                chunk.related_chunks.append(f"{doc_metadata['doc_id']}_pages_{i+1}")
        
        return chunks
    
    # =========================================================================
    # Spreadsheets : Par Feuille
    # =========================================================================
    
    def chunk_spreadsheet(self, file_path: Path, doc_metadata: Dict) -> List[ChunkMetadata]:
        """Chunk ODS/XLSX sémantique v2 : segmentation en zones + LLM Nemo.
        
        Stratégie :
        1. Extraire lignes brutes comme listes de cellules
        2. Segmenter en zones sémantiques (séparées par lignes vides)
        3. Sérialiser chaque zone en texte tabulaire lisible
        4. Appeler LLM Nemo pour convertir en texte naturel
        5. Découper si trop long
        """
        
        ext = file_path.suffix.lower()
        MAX_SHEETS = 20
        MAX_ROWS_PER_SHEET = 1000
        
        try:
            all_sheets_rows = self._extract_sheet_rows(file_path, ext, MAX_SHEETS, MAX_ROWS_PER_SHEET)
            
            # Initialiser LLM si pas encore fait
            if not hasattr(self, '_llm'):
                try:
                    from llm_provider import RAGConfig
                    config = RAGConfig()
                    self._llm = config.llm_provider
                except Exception as e:
                    logger.warning(f"LLM indisponible pour conversion tableur: {e}")
                    self._llm = None
            
            chunks = []
            for sheet_name, rows in all_sheets_rows:
                zones = self._segment_sheet_zones(rows)
                
                for zone_idx, zone_rows in enumerate(zones):
                    sub_zones = self._split_large_zone(zone_rows)
                    
                    for sub_zone in sub_zones:
                        raw_table_text = self._zone_to_table_text(sub_zone, sheet_name)
                        
                        if not raw_table_text or len(raw_table_text.split()) < 10:
                            continue
                        
                        llm_result = self._llm_convert_table(raw_table_text, sheet_name)
                        natural_text = llm_result['text']
                        tags = llm_result.get('tags', [])
                        
                        if not natural_text or len(natural_text.split()) < 15:
                            continue
                        
                        # Découper si trop long
                        sub_texts = self._split_natural_text(natural_text, max_words=500)
                    
                    for sub_text in sub_texts:
                        chunks.append(ChunkMetadata(
                            chunk_id=f"{doc_metadata['doc_id']}_sheet_{sheet_name}_{len(chunks)}",
                            parent_doc=doc_metadata['doc_id'],
                            chunk_text=sub_text,
                            chunk_type="sheet",
                            section=sheet_name,
                            position=f"{len(chunks)+1}/?",
                        ))
            
            # Update positions
            total = len(chunks)
            for i, chunk in enumerate(chunks):
                chunk.position = f"{i+1}/{total}"
                if i > 0:
                    chunk.related_chunks.append(chunks[i-1].chunk_id)
                if i < total - 1:
                    chunk.related_chunks.append(chunks[i+1].chunk_id)
            
            if not chunks:
                logger.warning(f"⚠️  Aucune feuille non-vide dans {file_path.name}")
                return [ChunkMetadata(
                    chunk_id=f"{doc_metadata['doc_id']}_empty",
                    parent_doc=doc_metadata['doc_id'],
                    chunk_text="[Spreadsheet vide]",
                    chunk_type="empty",
                    section="Vide",
                    position="1/1",
                )]
            
            logger.debug(f"  ✅ {len(chunks)} chunks sémantiques depuis {file_path.name}")
            return chunks
        
        except Exception as e:
            logger.error(f"❌ Erreur chunking spreadsheet: {e}")
            return [ChunkMetadata(
                chunk_id=f"{doc_metadata['doc_id']}_failed",
                parent_doc=doc_metadata['doc_id'],
                chunk_text=f"[Spreadsheet {file_path.suffix} - Extraction échouée: {str(e)}]",
                chunk_type="failed_extraction",
                section="Erreur",
                position="1/1",
            )]
    
    def _extract_sheet_rows(self, file_path: Path, ext: str,
                             max_sheets: int, max_rows: int) -> List[tuple]:
        """Extrait les lignes brutes de chaque feuille comme listes de cellules."""
        result = []
        
        if ext == '.ods':
            from odf import table as odf_table
            from odf.opendocument import load
            from odf import text as odf_text
            
            doc = load(str(file_path))
            sheets = doc.spreadsheet.getElementsByType(odf_table.Table)
            
            for i, sheet in enumerate(sheets[:max_sheets]):
                sheet_name = sheet.getAttribute('name') or f"Sheet{i+1}"
                rows = []
                for row in sheet.getElementsByType(odf_table.TableRow)[:max_rows]:
                    cells_el = row.getElementsByType(odf_table.TableCell)
                    cells = []
                    for cell in cells_el:
                        paras = cell.getElementsByType(odf_text.P)
                        cell_text = ' '.join(str(p) for p in paras).strip()
                        cells.append(cell_text)
                    if any(c for c in cells):
                        rows.append(cells)
                if rows:
                    result.append((sheet_name, rows))
        
        else:  # XLSX
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            sheet_names = wb.sheetnames[:max_sheets]
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                rows = []
                for row in list(sheet.iter_rows(values_only=True))[:max_rows]:
                    cells = [str(c).strip() if c is not None else '' for c in row]
                    if any(c for c in cells):
                        rows.append(cells)
                if rows:
                    result.append((sheet_name, rows))
            wb.close()
        
        return result
    
    def _segment_sheet_zones(self, rows: List[List[str]]) -> List[List[List[str]]]:
        """Segmente une feuille en zones sémantiques séparées par des lignes vides."""
        zones = []
        current_zone = []
        
        for row in rows:
            non_empty = [c for c in row if c.strip()]
            if not non_empty:
                if current_zone:
                    zones.append(current_zone)
                    current_zone = []
            else:
                current_zone.append(row)
        
        if current_zone:
            zones.append(current_zone)
        
        # Fusionner zones trop petites (1 ligne) avec la suivante
        merged = []
        buffer = None
        for zone in zones:
            if buffer is not None:
                merged.append(buffer + zone)
                buffer = None
            elif len(zone) == 1:
                max_cell = max((len(c) for c in zone[0] if c.strip()), default=0)
                if max_cell > 200:
                    merged.append(zone)
                else:
                    buffer = zone
            else:
                merged.append(zone)
        
        if buffer:
            if merged:
                merged[-1] = merged[-1] + buffer
            else:
                merged.append(buffer)
        
        return merged
    
    def _split_large_zone(self, zone_rows: List[List[str]], max_words: int = 250) -> List[List[List[str]]]:
        """Découpe une zone trop grande en sous-zones.
        
        Adapte dynamiquement le nombre de lignes par sous-zone pour
        que chaque sous-zone produise ~max_words mots en entrée LLM.
        La ligne d'en-têtes est préfixée à chaque sous-zone.
        """
        total_words = sum(len(' '.join(c for c in row if c.strip()).split()) for row in zone_rows)
        avg_words_per_row = max(total_words / max(len(zone_rows), 1), 1)
        max_rows = max(5, min(20, int(max_words / avg_words_per_row)))
        
        if len(zone_rows) <= max_rows:
            return [zone_rows]
        
        first_row = zone_rows[0]
        non_empty_first = [c for c in first_row if c.strip()]
        numeric_count = sum(1 for c in non_empty_first
                          if c.replace('.','').replace(',','').replace('-','').replace('/','').strip().isdigit())
        is_header = (
            len(non_empty_first) >= 2
            and all(len(c) < 80 for c in non_empty_first)
            and numeric_count < len(non_empty_first) * 0.5
        )
        
        header_row = zone_rows[0] if is_header else None
        data_start = 1 if is_header else 0
        data_rows = zone_rows[data_start:]
        
        sub_zones = []
        for i in range(0, len(data_rows), max_rows):
            batch = data_rows[i:i + max_rows]
            if header_row:
                sub_zones.append([header_row] + batch)
            else:
                sub_zones.append(batch)
        
        return sub_zones
    
    def _zone_to_table_text(self, zone_rows: List[List[str]], sheet_name: str) -> str:
        """Convertit une zone en texte tabulaire lisible pour le LLM."""
        lines = [f"Feuille : « {sheet_name} »"]
        
        for row in zone_rows:
            cells = [c.strip() for c in row]
            while cells and not cells[-1]:
                cells.pop()
            if not cells:
                continue
            display_cells = [c if c else '(vide)' for c in cells]
            lines.append(' | '.join(display_cells))
        
        return '\n'.join(lines)
    
    def _llm_convert_table(self, raw_table_text: str, sheet_name: str) -> dict:
        """Appelle Nemo LLM pour convertir un texte tabulaire en langage naturel + tags."""
        PROMPT = """Tu es un assistant spécialisé en protection des données personnelles (RGPD/CNIL).

TÂCHE : Convertis ce tableau en texte naturel, lisible et complet, puis catégorise-le.

RÈGLES POUR LE TEXTE :
- Écris des phrases complètes sujet-verbe-complément
- Conserve TOUTES les informations (noms, dates, durées, chiffres, articles de loi)
- Ne résume pas, ne paraphrase pas : restitue fidèlement le contenu
- Identifie les en-têtes de colonnes ou de lignes et utilise-les comme contexte
- Si le tableau a des en-têtes en première ligne OU en première colonne, associe chaque valeur à son en-tête
- Pour les listes de valeurs, utilise des énumérations naturelles
- Ne mets pas de titre, écris directement le texte
- Écris en français (même si le tableau est en anglais, traduis)

RÈGLES POUR LES TAGS :
- À la toute fin, ajoute une ligne commençant par [TAGS] suivie des thèmes RGPD couverts
- Choisis 1 à 3 tags parmi : droits des personnes, consentement, sécurité des données, durée de conservation, sous-traitance, base légale, données sensibles, transfert hors UE, cookies, violation de données, transparence, DPO, vidéosurveillance, finalité du traitement, registre des traitements, AIPD, anonymisation, minimisation, responsable de traitement, prospection commerciale, conformité RGPD, profilage, sanctions CNIL, données de santé, information des personnes
- Si aucun ne convient, propose un tag court et descriptif

TABLEAU :
{table_text}

TEXTE NATUREL :"""
        
        if self._llm is None:
            return self._mechanical_fallback(raw_table_text, sheet_name)
        
        try:
            prompt = PROMPT.format(table_text=raw_table_text)
            result = self._llm.generate(prompt, temperature=0.1, max_tokens=2000)
            result = result.strip()
            
            if 'TABLEAU :' in result or 'RÈGLES POUR' in result:
                return self._mechanical_fallback(raw_table_text, sheet_name)
            
            # Détection copie brute : si le LLM a recopié le texte avec les pipes
            pipe_count = result.count(' | ')
            input_pipe_count = raw_table_text.count(' | ')
            if input_pipe_count > 3 and pipe_count > input_pipe_count * 0.5:
                logger.warning(f"LLM a recopié le tableau ({pipe_count} pipes) pour {sheet_name}, retry simplifié")
                retry_prompt = (
                    "Réécris ce contenu en français courant, sans utiliser de tableaux ni de tirets. "
                    "Chaque information doit être une phrase complète. "
                    "Conserve toutes les données (noms, dates, durées, chiffres).\n\n"
                    f"{raw_table_text}\n\nTexte réécrit :"
                )
                try:
                    result = self._llm.generate(retry_prompt, temperature=0.2, max_tokens=2000).strip()
                    pipe_count = result.count(' | ')
                    if pipe_count > input_pipe_count * 0.3:
                        return self._mechanical_fallback(raw_table_text, sheet_name)
                except Exception as e:
                    logger.warning(f"Erreur retry LLM ({sheet_name}): {e}")
                    return self._mechanical_fallback(raw_table_text, sheet_name)
            
            # Extraire les tags
            tags = []
            text_lines = result.split('\n')
            clean_lines = []
            for line in text_lines:
                if line.strip().startswith('[TAGS]') or line.strip().startswith('[Tags]'):
                    tags = parse_tags(line)
                else:
                    clean_lines.append(line)
            
            clean_text = '\n'.join(clean_lines).strip()
            return {'text': f"Feuille « {sheet_name} »\n\n{clean_text}", 'tags': tags}
            
        except Exception as e:
            logger.warning(f"Erreur LLM conversion tableau ({sheet_name}): {e}")
            return self._mechanical_fallback(raw_table_text, sheet_name)
    
    def _mechanical_fallback(self, raw_table_text: str, sheet_name: str) -> dict:
        """Fallback mécanique si LLM indisponible : transformation propre sans pipes."""
        lines = raw_table_text.split('\n')
        text_parts = [f"Feuille « {sheet_name} »"]
        for line in lines:
            if line.startswith('Feuille :'):
                continue
            cells = [c.strip() for c in line.split(' | ')]
            cells = [c for c in cells if c and c != '(vide)']
            if not cells:
                continue
            if len(cells) == 1 and len(cells[0]) > 5:
                text_parts.append(cells[0])
            elif len(cells) == 2:
                text_parts.append(f"{cells[0]} : {cells[1]}.")
            elif len(cells) >= 3:
                values = ', '.join(cells[1:])
                text_parts.append(f"{cells[0]} : {values}.")
        return {'text': '\n'.join(text_parts), 'tags': []}
    
    def _split_natural_text(self, text: str, max_words: int = 500) -> List[str]:
        """Découpe un texte trop long en morceaux."""
        words = text.split()
        if len(words) <= max_words:
            return [text]
        
        paragraphs = text.split('\n\n')
        chunks = []
        current = []
        current_wc = 0
        
        for para in paragraphs:
            para_wc = len(para.split())
            if current_wc + para_wc > max_words and current:
                chunks.append('\n\n'.join(current))
                current = [para]
                current_wc = para_wc
            else:
                current.append(para)
                current_wc += para_wc
        
        if current:
            chunks.append('\n\n'.join(current))
        
        return chunks
    
    # =========================================================================
    # DOCX/ODT : Par Section
    # =========================================================================
    
    def chunk_document_text(self, file_path: Path, doc_metadata: Dict) -> List[ChunkMetadata]:
        """Chunk DOCX/ODT par sections (headings)"""
        
        ext = file_path.suffix.lower()
        
        try:
            if ext == '.docx':
                from docx import Document
                
                doc = Document(file_path)
                sections = []
                current_section = {'title': '', 'text': []}
                
                for para in doc.paragraphs:
                    if para.style.name.startswith('Heading'):
                        # Nouveau titre
                        if current_section['text']:
                            sections.append(current_section)
                        current_section = {'title': para.text, 'text': []}
                    else:
                        current_section['text'].append(para.text)
                
                # Dernier section
                if current_section['text']:
                    sections.append(current_section)
            
            elif ext == '.odt':
                from odf import text as odf_text
                from odf.opendocument import load
                
                doc = load(str(file_path))
                headings = doc.getElementsByType(odf_text.H)
                paragraphs = doc.getElementsByType(odf_text.P)
                
                # Simplification: tout en 1 chunk ou split par taille
                all_text = []
                for h in headings:
                    all_text.append(str(h))
                for p in paragraphs:
                    all_text.append(str(p))
                
                text = '\n'.join(all_text)
                return self._split_text_fixed(text, doc_metadata, "odt_content")
            
            # Créer chunks par section
            chunks = []
            for i, section in enumerate(sections):
                text = '\n'.join(section['text'])
                word_count = len(text.split())
                
                if word_count <= self.config.MAX_CHUNK_SIZE:
                    chunks.append(ChunkMetadata(
                        chunk_id=f"{doc_metadata['doc_id']}_section_{i}",
                        parent_doc=doc_metadata['doc_id'],
                        chunk_text=text,
                        chunk_type="semantic_section",
                        section=section['title'],
                        position=f"{i+1}/{len(sections)}",
                    ))
                else:
                    sub_chunks = self._split_text_fixed(text, doc_metadata, f"section_{i}", section['title'])
                    chunks.extend(sub_chunks)
            
            return chunks
        
        except Exception as e:
            logger.error(f"❌ Erreur chunking document text: {e}")
            # Retourner chunk d'erreur au lieu de fallback texte brut
            return [ChunkMetadata(
                chunk_id=f"{doc_metadata['doc_id']}_failed",
                parent_doc=doc_metadata['doc_id'],
                chunk_text=f"[Document {file_path.suffix} - Extraction échouée: {str(e)}]",
                chunk_type="failed_extraction",
                section="Erreur",
                position="1/1",
            )]
    
    # =========================================================================
    # Fallback : Split Fixe avec Overlap
    # =========================================================================
    
    def _split_text_fixed(self,
                         text: str,
                         doc_metadata: Dict,
                         prefix: str,
                         section: str = '') -> List[ChunkMetadata]:
        """Split texte en chunks de taille fixe avec overlap ET tracking des relations"""
        
        words = text.split()
        total_words = len(words)
        
        if total_words <= self.config.MAX_CHUNK_SIZE:
            return [ChunkMetadata(
                chunk_id=f"{doc_metadata['doc_id']}_{prefix}",
                parent_doc=doc_metadata['doc_id'],
                chunk_text=text,
                chunk_type="complete",
                section=section,
                position="1/1",
            )]
        
        # Split avec overlap
        chunks = []
        chunk_size = self.config.TARGET_CHUNK_SIZE
        overlap = self.config.OVERLAP
        
        i = 0
        while i < total_words:
            # Extraire chunk
            chunk_words = words[i:i+chunk_size]
            chunk_text = ' '.join(chunk_words)
            
            # Vérifier hard limit
            if len(chunk_words) > self.config.HARD_LIMIT_WORDS:
                logger.warning(f"  🔥 Chunk dépasse HARD_LIMIT ({len(chunk_words)} > {self.config.HARD_LIMIT_WORDS})")
                chunk_words = chunk_words[:self.config.HARD_LIMIT_WORDS]
                chunk_text = ' '.join(chunk_words)
            
            chunk_id = f"{doc_metadata['doc_id']}_{prefix}_chunk{len(chunks)}"
            
            # Calculer related_chunks (chunks adjacents via overlap)
            related = []
            if len(chunks) > 0:
                related.append(chunks[-1].chunk_id)  # Chunk précédent
            
            chunks.append(ChunkMetadata(
                chunk_id=chunk_id,
                parent_doc=doc_metadata['doc_id'],
                chunk_text=chunk_text,
                chunk_type="fixed_split",
                section=section,
                position=f"{len(chunks)+1}/?",
                is_continuation=len(chunks) > 0,
                related_chunks=related,
            ))
            
            # Avancer avec overlap
            i += chunk_size - overlap
        
        # Update positions ET ajouter chunk suivant dans related
        total = len(chunks)
        for i, chunk in enumerate(chunks):
            chunk.position = f"{i+1}/{total}"
            
            # Ajouter chunk suivant dans related_chunks
            if i < total - 1:
                chunk.related_chunks.append(chunks[i+1].chunk_id)
        
        return chunks
    
    # Note: chunk_fallback supprimé car lecture PDF comme texte brut = charabia
    # Utiliser plutôt chunk_text_from_file pour fichiers vraiment texte (rare)


# =========================================================================
# Tests
# =========================================================================

if __name__ == "__main__":
    import sys
    
    # Configuration
    config = ChunkingConfig()
    print(f"Configuration Chunking:")
    print(f"  MAX_CONTEXT_WINDOW: {config.MAX_CONTEXT_WINDOW} tokens")
    print(f"  MAX_CHUNK_TOKENS: {config.MAX_CHUNK_TOKENS} tokens")
    print(f"  HARD_LIMIT_WORDS: {config.HARD_LIMIT_WORDS} mots")
    print(f"  TARGET_CHUNK_SIZE: {config.TARGET_CHUNK_SIZE} mots")
    print(f"  MAX_CHUNK_SIZE: {config.MAX_CHUNK_SIZE} mots")
    
    # Test
    if len(sys.argv) > 1:
        test_file = Path(sys.argv[1])
        chunker = SemanticChunker(config)
        
        doc_meta = {'doc_id': test_file.stem}
        chunks = chunker.chunk_document(test_file, doc_meta)
        
        print(f"\n📊 Résultats chunking {test_file.name}:")
        print(f"   Chunks créés: {len(chunks)}")
        for chunk in chunks[:5]:
            print(f"   - {chunk.chunk_id}: {chunk.word_count} mots ({chunk.chunk_type})")
