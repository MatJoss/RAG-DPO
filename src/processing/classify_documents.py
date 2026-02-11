"""
Phase 5A : Pré-Catégorisation des Documents
Analyse chaque document pour déterminer son index primaire et secondaire
"""

import json
from pathlib import Path
import sys
import logging
from typing import Dict, List, Optional
from tqdm import tqdm
import time
from bs4 import BeautifulSoup

# Ajouter chemins
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root / 'src' / 'utils'))
sys.path.insert(0, str(project_root / 'src' / 'processing'))

from llm_provider import RAGConfig
from json_cleaner import clean_llm_json_response, safe_parse_json
from classification_validator import ClassificationValidator

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)
logging.getLogger("httpx").setLevel(logging.WARNING)


class DocumentClassifier:
    """Pré-catégorisation des documents (document-level)"""
    
    # Prompt de classification
    CLASSIFICATION_PROMPT = """Tu es un DPO senior classifiant des documents pour constituer une base de connaissances RAG professionnelle.

Tu dois déterminer la NATURE et les caractéristiques de chaque document pour permettre une recherche optimale.

## NATURES (choisis UNE SEULE) :

**DOCTRINE** — Interprétation juridique, cadre normatif, principes
  Exemples : lignes directrices CNIL, recommandations CEPD, analyses articles RGPD, FAQ sur les principes de licéité/finalité/minimisation, avis du G29/CEPD, délibérations interprétatives
  Signal clé : explique le POURQUOI juridique, interprète la loi

**GUIDE** — Mise en conformité opérationnelle, méthodologie, outils
  Exemples : guide AIPD pas-à-pas, modèle de registre, checklist sous-traitant, procédure de notification violation, template clause contractuelle, guide cookies, fiche pratique RH
  Signal clé : explique le COMMENT FAIRE, donne des étapes/modèles
  Attention : un guide qui cite des articles de loi reste un GUIDE

**SANCTION** — Décisions contentieuses, jurisprudence
  Exemples : délibération SAN-xxxx, amende, mise en demeure, clôture, avertissement
  Signal clé : code SAN/MED/CLOS, montant amende, décision contre un organisme identifié

**TECHNIQUE** — Sécurité IT, mesures techniques, cryptographie
  Exemples : guide sécurité des données, recommandations chiffrement, authentification, journalisation, pseudonymisation, architecture privacy by design
  Signal clé : détail technique (algorithme, protocole, configuration)

**MIXTE** — EXCEPTIONNEL (< 5%) : équilibre parfait doctrine + opérationnel
  Utiliser UNIQUEMENT si impossible de trancher après analyse

## RÈGLES DE DÉCISION :
- FAQ avec conseils pratiques → GUIDE
- Référentiel sectoriel (ex: santé, RH) → GUIDE (même s'il cite des obligations)
- Analyse juridique avec exemples illustratifs → DOCTRINE
- Recommandation technique CNIL (ex: mots de passe) → TECHNIQUE
- Document mixte juridique/pratique : regarde la MAJORITÉ du contenu

## IMPORTANCE (1-10) :
- 9-10 : Référence incontournable DPO (guide AIPD, sanctions majeures, recommandation CNIL clé)
- 7-8 : Très utile au quotidien (FAQ détaillée, fiche thématique complète)
- 5-6 : Utile ponctuellement (actualité sectorielle, cas d'usage spécifique)
- 3-4 : Information secondaire
- 1-2 : Marginal

Réponds en JSON STRICT :
{{
  "nature": "GUIDE",
  "secteurs": [],
  "importance": 8,
  "document_type": "checklist",
  "keywords": ["conformité", "RGPD"],
  "raison": "Checklist pratique pour mise en conformité"
}}

- "nature" : DOCTRINE | GUIDE | SANCTION | TECHNIQUE | MIXTE
- "secteurs" : [] si générique, ou ["Santé"] ["RH"] ["Marketing"] ["Éducation"] ["Vidéosurveillance"] ["Banque"] etc.
- "document_type" : guide | faq | referentiel | deliberation | recommandation | fiche | modele | article | autre
"""
    
    def __init__(self, project_root: str = '.', fresh: bool = False):
        self.project_root = Path(project_root)
        self.data_path = self.project_root / 'data'
        self.cnil_path = self.data_path / 'raw' / 'cnil'
        self.keep_dir = self.data_path / 'keep' / 'cnil'
        
        # Fichiers
        self.manifest_file = self.cnil_path / 'keep_manifest.json'
        self.cache_file = self.cnil_path / 'document_classification_cache.json'
        self.output_file = self.cnil_path / 'document_metadata.json'
        
        # Mode fresh : purger cache et output existants
        if fresh:
            for f in [self.cache_file, self.output_file]:
                if f.exists():
                    f.unlink()
                    logger.info(f"🗑️  Supprimé (fresh) : {f.name}")
        
        # Cache
        self.cache = self._load_cache()
        
        # Validator
        self.validator = ClassificationValidator()
        
        # LLM
        try:
            config = RAGConfig()
            self.llm = config.llm_provider
            self.mode = config.mode
            logger.info(f"🤖 LLM initialisé en mode : {self.mode}")
        except Exception as e:
            logger.error(f"❌ Erreur init LLM : {e}")
            raise
        
        # Stats
        self.stats = {
            'total_documents': 0,
            'html_processed': 0,
            'pdf_processed': 0,
            'docs_processed': 0,
            'images_processed': 0,
            'cached': 0,
            'errors': 0,
            'json_errors': 0,
            'llm_errors': 0,
            'by_primary_index': {
                'FONDAMENTAUX': 0,
                'OPERATIONNEL': 0,
                'SANCTIONS': 0,
                'SECTORIELS': 0,
                'TECHNIQUE': 0,
            },
            'by_nature': {
                'DOCTRINE': 0,
                'GUIDE': 0,
                'SANCTION': 0,
                'TECHNIQUE': 0,
                'MIXTE': 0,
            }
        }
    
    def _load_cache(self) -> Dict:
        """Charge le cache"""
        if self.cache_file.exists():
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    cache = json.load(f)
                logger.info(f"📦 Cache chargé : {len(cache)} documents")
                return cache
            except:
                return {}
        return {}
    
    def _save_cache(self):
        """Sauvegarde le cache"""
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"❌ Erreur sauvegarde cache : {e}")
    
    def extract_preview_html(self, file_path: Path, max_words: int = 800) -> str:
        """Extrait preview d'un HTML en ciblant le contenu principal"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                soup = BeautifulSoup(f, 'lxml')
            
            # Extraire titre
            title = ''
            title_tag = soup.find(['h1', 'title'])
            if title_tag:
                title = title_tag.get_text(strip=True)
            
            # Cibler le bloc de contenu principal (structure CNIL)
            content_block = (
                soup.find(class_='region-content')
                or soup.find('main')
                or soup.find('article')
                or soup.find(class_='field-name-body')
                or soup  # Fallback
            )
            
            # Supprimer bruit
            for tag in content_block(['script', 'style', 'nav', 'header', 'footer', 
                                       'aside', 'iframe', 'noscript', 'svg']):
                tag.decompose()
            
            # Supprimer menus, breadcrumbs, pagination
            for nav_block in content_block.find_all(class_=lambda c: c and any(
                x in str(c).lower() for x in [
                    'menu-push', 'breadcrumb', 'pager', 'pagination',
                    'nav-', 'share-', 'social', 'cookie', 'back-to-top'
                ]
            )):
                nav_block.decompose()
            
            text = content_block.get_text(separator=' ', strip=True)
            
            # Limiter
            words = text.split()[:max_words]
            preview = ' '.join(words)
            
            return f"Titre: {title}\n\n{preview}"
        
        except Exception as e:
            logger.debug(f"⚠️  Erreur extraction HTML: {e}")
            return "[HTML - Extraction échouée]"
    
    def extract_preview_pdf(self, file_path: Path, max_words: int = 800) -> str:
        """Extrait preview d'un PDF (multi-méthodes robuste avec OCR)"""
        
        # Méthode 1: PyMuPDF
        try:
            import fitz
            doc = fitz.open(file_path)
            
            # Extraire premières pages
            text_parts = []
            for i in range(min(3, len(doc))):  # 3 premières pages max
                page_text = doc[i].get_text()
                text_parts.append(page_text)
            
            doc.close()
            
            text = '\n'.join(text_parts)
            words = text.split()[:max_words]
            preview = ' '.join(words)
            
            if len(preview) > 100:
                logger.debug(f"  ✅ PyMuPDF: {file_path.name}")
                return preview
        
        except Exception as e:
            logger.debug(f"  ⚠️  PyMuPDF échec: {e}")
        
        # Méthode 2: pdfplumber
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                text_parts = []
                for i in range(min(3, len(pdf.pages))):
                    page_text = pdf.pages[i].extract_text()
                    if page_text:
                        text_parts.append(page_text)
                
                text = '\n'.join(text_parts)
                words = text.split()[:max_words]
                preview = ' '.join(words)
                
                if len(preview) > 100:
                    logger.debug(f"  ✅ pdfplumber: {file_path.name}")
                    return preview
        
        except Exception as e:
            logger.debug(f"  ⚠️  pdfplumber échec: {e}")
        
        # Méthode 3: PyPDF2
        try:
            import PyPDF2
            with open(file_path, 'rb') as f:
                pdf = PyPDF2.PdfReader(f)
                text_parts = []
                for i in range(min(3, len(pdf.pages))):
                    page_text = pdf.pages[i].extract_text()
                    if page_text:
                        text_parts.append(page_text)
                
                text = '\n'.join(text_parts)
                words = text.split()[:max_words]
                preview = ' '.join(words)
                
                if len(preview) > 100:
                    logger.debug(f"  ✅ PyPDF2: {file_path.name}")
                    return preview
        
        except Exception as e:
            logger.debug(f"  ⚠️  PyPDF2 échec: {e}")
        
        # Méthode 4: OCR Tesseract (PDFs scannés)
        try:
            import fitz
            from PIL import Image
            import pytesseract
            import io
            
            logger.info(f"  🔍 OCR Tesseract pour preview: {file_path.name}")
            
            doc = fitz.open(file_path)
            text_parts = []
            
            # OCR seulement sur 2 premières pages (preview rapide)
            for page_num in range(min(2, len(doc))):
                page = doc[page_num]
                
                # Convertir en image
                pix = page.get_pixmap(dpi=150)
                img_data = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_data))
                
                # OCR
                page_text = pytesseract.image_to_string(img, lang='fra')
                
                if page_text and len(page_text.strip()) > 50:
                    text_parts.append(page_text)
            
            doc.close()
            
            if text_parts:
                text = '\n'.join(text_parts)
                words = text.split()[:max_words]
                preview = ' '.join(words)
                
                if len(preview) > 100:
                    logger.debug(f"  ✅ OCR Tesseract: {file_path.name}")
                    return preview
        
        except ImportError:
            logger.debug(f"  ⚠️  Tesseract non installé")
        except Exception as e:
            logger.debug(f"  ⚠️  OCR échec: {e}")
        
        # Échec total
        logger.warning(f"⚠️  Impossible d'extraire preview de {file_path.name}")
        return "[PDF - Extraction échouée avec toutes les méthodes (PyMuPDF, pdfplumber, PyPDF2, OCR)]"
    
    def extract_preview_doc(self, file_path: Path, max_words: int = 800) -> str:
        """Extrait preview d'un ODS/DOCX/ODT (toutes les feuilles pour spreadsheets)"""
        ext = file_path.suffix.lower()
        
        # Limites de sécurité (cohérent avec chunking_strategy)
        MAX_SHEETS = 20
        MAX_ROWS_PER_SHEET = 50  # Pour preview, moins de lignes que chunking
        
        try:
            if ext == '.ods':
                from odf import table as odf_table, text as odf_text
                from odf.opendocument import load
                
                doc = load(str(file_path))
                sheets = doc.spreadsheet.getElementsByType(odf_table.Table)
                
                # Traiter TOUTES les feuilles (max 20)
                sheets_to_process = sheets[:MAX_SHEETS]
                
                if len(sheets) > MAX_SHEETS:
                    logger.debug(f"  ⚠️  ODS a {len(sheets)} feuilles, preview limité à {MAX_SHEETS}")
                
                all_text_parts = []
                
                for i, sheet in enumerate(sheets_to_process):
                    sheet_name = sheet.getAttribute('name') or f"Sheet{i+1}"
                    all_text_parts.append(f"[Feuille: {sheet_name}]")
                    
                    rows = sheet.getElementsByType(odf_table.TableRow)
                    for row in rows[:MAX_ROWS_PER_SHEET]:
                        cells = row.getElementsByType(odf_table.TableCell)
                        for cell in cells:
                            paras = cell.getElementsByType(odf_text.P)
                            cell_text = ' '.join([str(p) for p in paras])
                            if cell_text.strip():
                                all_text_parts.append(cell_text.strip())
                
                text = ' '.join(all_text_parts)
                words = text.split()[:max_words]
                return ' '.join(words)
            
            elif ext == '.xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                # Traiter TOUTES les feuilles (max 20)
                sheet_names = wb.sheetnames[:MAX_SHEETS]
                
                if len(wb.sheetnames) > MAX_SHEETS:
                    logger.debug(f"  ⚠️  XLSX a {len(wb.sheetnames)} feuilles, preview limité à {MAX_SHEETS}")
                
                all_text_parts = []
                
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    all_text_parts.append(f"[Feuille: {sheet_name}]")
                    
                    for row in list(sheet.iter_rows(values_only=True))[:MAX_ROWS_PER_SHEET]:
                        row_text = ' '.join([str(cell) for cell in row if cell])
                        if row_text.strip():
                            all_text_parts.append(row_text)
                
                wb.close()
                
                text = ' '.join(all_text_parts)
                words = text.split()[:max_words]
                return ' '.join(words)
            
            elif ext in ['.docx', '.odt']:
                if ext == '.docx':
                    from docx import Document
                    doc = Document(file_path)
                    paras = [p.text for p in doc.paragraphs[:30]]
                    text = '\n'.join(paras)
                else:  # .odt
                    from odf import text as odf_text
                    from odf.opendocument import load
                    doc = load(str(file_path))
                    paras = doc.getElementsByType(odf_text.P)
                    text = '\n'.join([str(p) for p in paras[:30]])
                
                words = text.split()[:max_words]
                return ' '.join(words)
        
        except Exception as e:
            logger.debug(f"⚠️  Erreur extraction {file_path.name}: {e}")
            return f"[{ext[1:].upper()} - Extraction échouée: {str(e)}]"
        
        return f"[{ext[1:].upper()}]"
    
    def classify_document(self, doc_info: Dict) -> Dict:
        """Classifie un document avec le LLM (ou déterministe pour images)"""
        
        # Cache
        cache_key = doc_info['file_path']
        if cache_key in self.cache:
            result = self.cache[cache_key].copy()
            result['cached'] = True
            self.stats['cached'] += 1
            return result
        
        # Images SCHEMA_DPO → classification déterministe (pas de LLM)
        if doc_info.get('type') == 'image':
            return self._classify_image_deterministic(doc_info)
        
        # Construire prompt
        file_path = self.project_root / doc_info['file_path']
        file_type = file_path.suffix.lower()[1:]
        
        # Extraire preview selon type
        if file_type == 'html':
            preview = self.extract_preview_html(file_path)
        elif file_type == 'pdf':
            preview = self.extract_preview_pdf(file_path)
        elif file_type in ['ods', 'xlsx', 'docx', 'odt']:
            preview = self.extract_preview_doc(file_path)
        else:
            preview = "[Type de fichier non supporté]"
        
        # Titre depuis metadata si dispo
        title = doc_info.get('title', file_path.stem)
        
        # Contexte enrichi avec relations
        context_parts = []
        
        # Parent URL (pour PDFs/Docs)
        if doc_info.get('parent_url'):
            context_parts.append(f"Rattaché à la page : {doc_info['parent_url']}")
        
        # Ressources liées (pour HTML)
        if doc_info.get('related_resources'):
            resources = doc_info['related_resources']
            if resources:
                types = {}
                for res in resources:
                    res_type = res.get('type', 'unknown')
                    types[res_type] = types.get(res_type, 0) + 1
                
                res_desc = ', '.join([f"{count} {type}(s)" for type, count in types.items()])
                context_parts.append(f"Avec ressources attachées : {res_desc}")
        
        context_info = "\n".join(context_parts) if context_parts else "Aucune relation détectée"
        
        user_prompt = f"""Document : {title}
Type : {file_type.upper()}
URL source : {doc_info.get('url', 'N/A')}

Contexte relationnel :
{context_info}

Preview du contenu :
{preview}

Catégorise ce document."""

        full_prompt = f"{self.CLASSIFICATION_PROMPT}\n\n{user_prompt}"
        
        try:
            response = self.llm.generate(full_prompt, temperature=0.0, max_tokens=500)
            
            # Nettoyage JSON avec module dédié
            response_clean = clean_llm_json_response(response)
            
            # Parsing JSON sécurisé
            result = safe_parse_json(response_clean)
            
            # Validation minimale : le LLM doit retourner au moins "nature"
            if not result.get('nature'):
                logger.warning(f"⚠️  Réponse LLM sans 'nature' pour {title[:50]} → classification par défaut")
                self.stats['llm_errors'] += 1
                self.stats['errors'] += 1
                return self._default_classification(doc_info, file_type, title)
            
            # MAPPING DÉTERMINISTE : nature → index (Solution ChatGPT)
            result = self._map_nature_to_index(result)
            
            # VALIDATION POST-CLASSIFICATION (désactivable si besoin)
            # result = self.validator.validate(result, preview, title)
            
            result['cached'] = False
            result['file_type'] = file_type
            result['title'] = title
            
            # Ajouter contexte relationnel au résultat
            result['parent_url'] = doc_info.get('parent_url')
            result['has_related_resources'] = bool(doc_info.get('related_resources'))
            if doc_info.get('related_resources'):
                result['related_resources_count'] = len(doc_info['related_resources'])
            
            # Stats (on utilise maintenant la nature pour les stats)
            nature = result.get('nature', 'GUIDE')
            primary = result.get('primary_index', 'OPERATIONNEL')
            
            if primary in self.stats['by_primary_index']:
                self.stats['by_primary_index'][primary] += 1
            
            if nature in self.stats['by_nature']:
                self.stats['by_nature'][nature] += 1
            
            # Log avec la NATURE (plus pertinent maintenant)
            logger.info(f"✅ {nature:12s} → {primary:12s} (imp:{result.get('importance', 5)}) - {title[:50]}...")
            
            # Cache
            self.cache[cache_key] = result
            
            time.sleep(0.5 if self.mode == 'local' else 1.0)
            
            return result
        
        except json.JSONDecodeError as e:
            # DUMP du JSON problématique pour debug
            debug_dir = self.data_path / 'debug_json_errors'
            debug_dir.mkdir(exist_ok=True)
            
            # Nom fichier = hash du file_path + extension
            file_name = Path(file_path).stem if file_path else 'unknown'
            debug_file = debug_dir / f"{file_name}_{file_type}.txt"
            
            # Sauvegarder le JSON problématique avec contexte
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write("=" * 70 + "\n")
                f.write("JSON PARSING ERROR DEBUG\n")
                f.write("=" * 70 + "\n\n")
                f.write(f"Fichier source : {file_path}\n")
                f.write(f"Type : {file_type}\n")
                f.write(f"Titre : {title}\n\n")
                f.write(f"Erreur JSON : {str(e)}\n")
                f.write(f"Position : line {e.lineno} column {e.colno}\n\n")
                f.write("=" * 70 + "\n")
                f.write("RÉPONSE BRUTE DU LLM\n")
                f.write("=" * 70 + "\n\n")
                f.write(response + "\n\n")
                f.write("=" * 70 + "\n")
                f.write("APRÈS NETTOYAGE (ce qui a été tenté de parser)\n")
                f.write("=" * 70 + "\n\n")
                # Pretty print avec indentation pour lisibilité
                try:
                    # Essayer d'indenter même si invalide
                    lines = response_clean.split('\n')
                    for i, line in enumerate(lines, 1):
                        f.write(f"{i:3d} | {line}\n")
                except:
                    f.write(response_clean)
                
                f.write("\n\n")
                f.write("=" * 70 + "\n")
                f.write("LIGNE PROBLÉMATIQUE\n")
                f.write("=" * 70 + "\n\n")
                try:
                    lines = response_clean.split('\n')
                    if e.lineno <= len(lines):
                        f.write(f"Ligne {e.lineno}: {lines[e.lineno-1]}\n")
                        f.write(" " * (e.colno + 10) + "^\n")
                except:
                    pass
            
            logger.error(f"❌ JSON invalide pour {title}: {e}")
            logger.error(f"   💾 JSON dumped to: {debug_file}")
            
            self.stats['json_errors'] += 1
            self.stats['errors'] += 1
            return self._default_classification(doc_info, file_type, title)
        
        except Exception as e:
            # DUMP debug (même logique que pour json.JSONDecodeError)
            debug_dir = self.data_path / 'debug_json_errors'
            debug_dir.mkdir(exist_ok=True)
            file_name = Path(file_path).stem if file_path else 'unknown'
            debug_file = debug_dir / f"{file_name}_{file_type}_llm.txt"
            
            try:
                with open(debug_file, 'w', encoding='utf-8') as f:
                    f.write("=" * 70 + "\n")
                    f.write("LLM/PARSING ERROR DEBUG\n")
                    f.write("=" * 70 + "\n\n")
                    f.write(f"Fichier source : {file_path}\n")
                    f.write(f"Type : {file_type}\n")
                    f.write(f"Titre : {title}\n")
                    f.write(f"Erreur : {type(e).__name__}: {e}\n\n")
                    f.write("=" * 70 + "\n")
                    f.write("RÉPONSE BRUTE DU LLM\n")
                    f.write("=" * 70 + "\n\n")
                    f.write(repr(response) + "\n\n")
                    f.write("=" * 70 + "\n")
                    f.write("APRÈS NETTOYAGE\n")
                    f.write("=" * 70 + "\n\n")
                    f.write(repr(response_clean) + "\n")
            except Exception:
                pass
            
            logger.error(f"❌ Erreur classification {title}: {e}")
            logger.error(f"   💾 Debug dumped to: {debug_file}")
            self.stats['llm_errors'] += 1
            self.stats['errors'] += 1
            return self._default_classification(doc_info, file_type, title)
    
    def _map_nature_to_index(self, result: Dict) -> Dict:
        """Mapping déterministe : nature juridique → index (Solution ChatGPT)
        
        Args:
            result: Classification LLM avec champ "nature"
            
        Returns:
            Classification avec primary_index et secondary_indexes dérivés
        """
        
        nature = result.get('nature', 'GUIDE')
        
        # Mapping déterministe
        nature_to_index = {
            'DOCTRINE': {
                'primary': 'FONDAMENTAUX',
                'secondary': []
            },
            'GUIDE': {
                'primary': 'OPERATIONNEL',
                'secondary': []
            },
            'SANCTION': {
                'primary': 'SANCTIONS',
                'secondary': []
            },
            'TECHNIQUE': {
                'primary': 'TECHNIQUE',
                'secondary': []
            },
            'MIXTE': {
                'primary': 'FONDAMENTAUX',
                'secondary': ['OPERATIONNEL']  # Doctrine + conseils pratiques
            }
        }
        
        # Appliquer mapping
        mapping = nature_to_index.get(nature, {'primary': 'OPERATIONNEL', 'secondary': []})
        result['primary_index'] = mapping['primary']
        result['secondary_indexes'] = mapping['secondary']
        
        # Si secteurs spécifiés ET doc sectoriel → ajouter SECTORIELS en secondaire
        secteurs = result.get('secteurs', [])
        if secteurs and len(secteurs) <= 2:
            # Document vraiment sectoriel
            if 'SECTORIELS' not in result['secondary_indexes']:
                result['secondary_indexes'].append('SECTORIELS')
        
        logger.debug(f"  🔄 Mapping: {nature} → {result['primary_index']} (secondaires: {result['secondary_indexes']})")
        
        return result
    
    def _default_classification(self, doc_info: Dict, file_type: str, title: str) -> Dict:
        """Classification par défaut en cas d'erreur"""
        return {
            "nature": "GUIDE",
            "primary_index": "OPERATIONNEL",
            "secondary_indexes": [],
            "secteurs": [],
            "importance": 5,
            "document_type": "unknown",
            "keywords": [],
            "raison": "Classification par defaut (erreur LLM)",
            "cached": False,
            "file_type": file_type,
            "title": title,
            "error": True
        }
    
    def _classify_image_deterministic(self, doc_info: Dict) -> Dict:
        """Classification déterministe pour images SCHEMA_DPO (pas de LLM).
        
        Les images ont déjà été analysées par Phase 4B (OCR + LLaVA).
        On mappe directement leur classification en metadata document.
        """
        img_cls = doc_info.get('image_classification', {})
        title = doc_info.get('title', '')
        
        result = {
            'nature': 'TECHNIQUE',
            'primary_index': 'TECHNIQUE',
            'secondary_indexes': [],
            'secteurs': [],
            'importance': 6,  # Schémas = assez importants (visuels pédagogiques)
            'document_type': 'schema_image',
            'keywords': ['schéma', 'diagramme', 'visuel', 'DPO', 'RGPD'],
            'raison': img_cls.get('description', '') or title,
            'cached': False,
            'file_type': 'image',
            'title': title,
            'parent_url': doc_info.get('parent_url'),
            'has_related_resources': False,
            'image_ocr_words': img_cls.get('ocr_words', 0),
            'image_method': 'deterministic_from_phase4b',
        }
        
        # Stats
        if 'TECHNIQUE' in self.stats['by_primary_index']:
            self.stats['by_primary_index']['TECHNIQUE'] += 1
        if 'TECHNIQUE' in self.stats['by_nature']:
            self.stats['by_nature']['TECHNIQUE'] += 1
        
        logger.info(f"✅ {'TECHNIQUE':12s} → {'TECHNIQUE':12s} (img, déterministe) - {title[:50]}")
        
        # Cache
        self.cache[doc_info['file_path']] = result
        
        return result
    
    def run(self, max_documents: Optional[int] = None):
        """Exécute la classification complète"""
        
        print("=" * 70)
        print("📋 PHASE 5A : PRÉ-CATÉGORISATION DES DOCUMENTS")
        print("=" * 70)
        
        # Charger manifest
        if not self.manifest_file.exists():
            print("\n❌ keep_manifest.json introuvable")
            print("   Lancez organize_keep_archive.py d'abord")
            return
        
        with open(self.manifest_file, 'r', encoding='utf-8') as f:
            manifest = json.load(f)
        
        # Collecter tous les documents avec leurs relations
        all_docs = []
        relationships = manifest.get('relationships', {})
        
        # HTML (liste d'objets avec 'url', 'file', 'metadata')
        for item in manifest.get('html', []):
            url = item['url']
            
            # Récupérer ressources liées depuis relationships
            related_resources = relationships.get(url, [])
            
            all_docs.append({
                'url': url,
                'file_path': item['metadata']['file_path'],
                'type': 'html',
                'title': item.get('metadata', {}).get('title', ''),
                'parent_url': None,  # HTML n'a pas de parent
                'related_resources': related_resources,  # PDFs/docs attachés
            })
        
        # PDFs (liste d'objets)
        for item in manifest.get('pdfs', []):
            all_docs.append({
                'url': item['url'],
                'file_path': item['metadata']['file_path'],
                'type': 'pdf',
                'title': '',
                'parent_url': item.get('parent_url'),  # Page HTML parente
                'related_resources': [],
            })
        
        # Docs (liste d'objets)
        for item in manifest.get('docs', []):
            all_docs.append({
                'url': item['url'],
                'file_path': item['metadata']['file_path'],
                'type': 'doc',
                'title': '',
                'parent_url': item.get('parent_url'),  # Page HTML parente
                'related_resources': [],
            })
        
        # Images SCHEMA_DPO (déjà classifiées par Phase 4B)
        for item in manifest.get('images', []):
            img_cls = item.get('classification', {})
            all_docs.append({
                'url': item['url'],
                'file_path': item['metadata']['file_path'],
                'type': 'image',
                'title': img_cls.get('description', '') or f"Schéma DPO - {Path(item['file']).stem}",
                'parent_url': item.get('parent_url'),
                'related_resources': [],
                'image_classification': img_cls,  # OCR + LLaVA data from Phase 4B
            })
        
        self.stats['total_documents'] = len(all_docs)
        
        # Limiter si test
        if max_documents:
            all_docs = all_docs[:max_documents]
            print(f"\n🧪 MODE TEST : {max_documents} documents maximum")
        
        print(f"\n📊 Documents à classifier : {len(all_docs)}")
        print(f"   HTML   : {len(manifest.get('html', []))}")
        print(f"   PDFs   : {len(manifest.get('pdfs', []))}")
        print(f"   Docs   : {len(manifest.get('docs', []))}")
        print(f"   Images : {len(manifest.get('images', []))} (SCHEMA_DPO, classification déterministe)")        # Statistiques sur relations
        html_with_resources = sum(1 for item in manifest.get('html', []) 
                                  if relationships.get(item['url'], []))
        pdfs_with_parent = sum(1 for item in manifest.get('pdfs', []) 
                              if item.get('parent_url'))
        
        print(f"\n🔗 Relations détectées :")
        print(f"   HTML avec ressources : {html_with_resources}")
        print(f"   PDFs avec parent     : {pdfs_with_parent}")
        
        # Estimation (2s par doc non-caché en moyenne)
        uncached = sum(1 for d in all_docs if d['file_path'] not in self.cache)
        est_min = uncached * 2 / 60
        print(f"\n⏱️  À traiter : {uncached} (cache: {len(all_docs) - uncached})")
        print(f"   Durée estimée : ~{est_min:.0f} minutes ({est_min/60:.1f}h)")
        
        if not max_documents and sys.stdin.isatty():
            input("\n   Appuyez sur Entrée pour continuer...")
        
        # Classification
        results = {}
        save_counter = 0
        
        print(f"\n🔄 Classification en cours...")
        
        for doc in tqdm(all_docs, desc="Classification"):
            classification = self.classify_document(doc)
            
            results[doc['file_path']] = classification
            
            # Type tracking
            if doc['type'] == 'html':
                self.stats['html_processed'] += 1
            elif doc['type'] == 'pdf':
                self.stats['pdf_processed'] += 1
            elif doc['type'] == 'image':
                self.stats['images_processed'] += 1
            else:
                self.stats['docs_processed'] += 1
            
            # Sauvegarder cache tous les 20
            save_counter += 1
            if save_counter % 20 == 0:
                self._save_cache()
        
        # Sauvegarde finale
        self._save_cache()
        self._save_results(results)
        
        # Résumé
        self._print_summary()
    
    def _save_results(self, results: Dict):
        """Sauvegarde les résultats"""
        output = {
            'metadata': results,
            'stats': self.stats,
            'timestamp': time.strftime('%Y-%m-%dT%H:%M:%S')
        }
        
        with open(self.output_file, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        logger.info(f"💾 Résultats : {self.output_file}")
    
    def _print_summary(self):
        """Affiche le résumé"""
        print("\n" + "=" * 70)
        print("📊 RÉSUMÉ - PRÉ-CATÉGORISATION")
        print("=" * 70)
        
        print(f"\n📄 Documents traités :")
        print(f"   HTML   : {self.stats['html_processed']}")
        print(f"   PDFs   : {self.stats['pdf_processed']}")
        print(f"   Docs   : {self.stats['docs_processed']}")
        print(f"   Images : {self.stats['images_processed']} (SCHEMA_DPO, déterministe)")
        print(f"   TOTAL  : {self.stats['total_documents']}")
        
        print(f"\n🔬 Classification par NATURE juridique :")
        for nature, count in sorted(self.stats['by_nature'].items()):
            pct = count / max(1, self.stats['total_documents']) * 100
            print(f"   {nature:15s} : {count:4d} ({pct:5.1f}%)")
        
        print(f"\n📂 Répartition finale par INDEX thématique :")
        for index, count in sorted(self.stats['by_primary_index'].items()):
            pct = count / max(1, self.stats['total_documents']) * 100
            print(f"   {index:15s} : {count:4d} ({pct:5.1f}%)")
        
        print(f"\n💡 Note : Les index sont déduits automatiquement des natures")
        print(f"   DOCTRINE → FONDAMENTAUX")
        print(f"   GUIDE → OPERATIONNEL")
        print(f"   MIXTE → FONDAMENTAUX (primaire) + OPERATIONNEL (secondaire)")
        
        # Stats succès/erreurs
        total_processed = self.stats['total_documents']
        success = total_processed - self.stats['errors']
        success_rate = (success / max(1, total_processed)) * 100
        
        print(f"\n✅ Taux de succès :")
        print(f"   Réussies       : {success}/{total_processed} ({success_rate:.1f}%)")
        if self.stats['errors'] > 0:
            print(f"   ⚠️  Erreurs JSON  : {self.stats['json_errors']}")
            print(f"   ⚠️  Erreurs LLM   : {self.stats['llm_errors']}")
            print(f"   ⚠️  Total erreurs : {self.stats['errors']} ({self.stats['errors']/max(1,total_processed)*100:.1f}%)")
        
        print(f"\n💾 Cache : {self.stats['cached']}/{self.stats['total_documents']}")
        
        print("\n" + "=" * 70)
        print(f"💾 Résultats : {self.output_file}")
        print(f"💾 Cache : {self.cache_file}")
        print("=" * 70)


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Phase 5A - Pré-catégorisation documents')
    parser.add_argument('--project-root', type=str, default='.', help='Racine du projet')
    parser.add_argument('--test', type=int, help='Tester sur N documents')
    parser.add_argument('--verbose', '-v', action='store_true', help='Mode verbose')
    parser.add_argument('--fresh', action='store_true', help='Ignorer cache existant, tout reclassifier')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger("__main__").setLevel(logging.DEBUG)
    
    classifier = DocumentClassifier(args.project_root, fresh=args.fresh)
    classifier.run(max_documents=args.test)


if __name__ == "__main__":
    main()