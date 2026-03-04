"""
Chunking Sémantique
Découpage intelligent basé sur la structure des documents
"""

import sys
import re
from pathlib import Path
from typing import List, Dict, Tuple
from bs4 import BeautifulSoup
import logging

# Ajouter chemin utils
_project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(_project_root / 'src' / 'utils'))
from rgpd_topics import parse_tags, TAG_PROMPT_TABLE

logger = logging.getLogger(__name__)


class SemanticChunker:
    """Chunker sémantique basé sur la structure des documents"""
    
    # Tailles cibles
    MIN_CHUNK_SIZE = 200   # mots
    TARGET_CHUNK_SIZE = 500  # mots
    MAX_CHUNK_SIZE = 1000  # mots
    
    def chunk_html(self, file_path: Path) -> List[Dict]:
        """Chunke un HTML sur les titres h2/h3"""
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                soup = BeautifulSoup(f, 'lxml')
            
            # Supprimer bruit
            for tag in soup(['script', 'style', 'nav', 'header', 'footer', 'aside']):
                tag.decompose()
            
            # Extraire titre principal
            title = ''
            title_tag = soup.find(['h1', 'title'])
            if title_tag:
                title = title_tag.get_text(strip=True)
            
            # Découper sur h2
            chunks = []
            main = soup.find(['main', 'article', 'div']) or soup.body or soup
            
            current_chunk = {
                'heading': title,
                'content': [],
                'level': 1
            }
            
            for element in main.descendants:
                if element.name in ['h2', 'h3', 'h4']:
                    # Sauvegarder chunk précédent si non vide
                    if current_chunk['content']:
                        chunk_text = ' '.join(current_chunk['content'])
                        if len(chunk_text.split()) >= self.MIN_CHUNK_SIZE:
                            chunks.append({
                                'text': chunk_text,
                                'heading': current_chunk['heading'],
                                'type': 'section'
                            })
                    
                    # Nouveau chunk
                    current_chunk = {
                        'heading': element.get_text(strip=True),
                        'content': [],
                        'level': int(element.name[1])
                    }
                
                elif element.name in ['p', 'li', 'td']:
                    text = element.get_text(strip=True)
                    if text:
                        current_chunk['content'].append(text)
            
            # Dernier chunk
            if current_chunk['content']:
                chunk_text = ' '.join(current_chunk['content'])
                if len(chunk_text.split()) >= self.MIN_CHUNK_SIZE:
                    chunks.append({
                        'text': chunk_text,
                        'heading': current_chunk['heading'],
                        'type': 'section'
                    })
            
            # Si pas de chunks (page simple), prendre tout
            if not chunks:
                full_text = main.get_text(separator=' ', strip=True)
                chunks = self._split_by_size(full_text, title)
            
            logger.debug(f"  HTML chunked: {len(chunks)} chunks from {file_path.name}")
            return chunks
        
        except Exception as e:
            logger.error(f"  Erreur chunking HTML {file_path.name}: {e}")
            return []
    
    def chunk_pdf(self, file_path: Path) -> List[Dict]:
        """Chunke un PDF sur TOC + structure"""
        
        try:
            import fitz
            doc = fitz.open(file_path)
            
            # Essayer d'extraire TOC
            toc = doc.get_toc()
            
            if toc and len(toc) > 2:
                # Chunker sur TOC
                chunks = self._chunk_pdf_with_toc(doc, toc)
            else:
                # Chunker sur pages + font changes
                chunks = self._chunk_pdf_structural(doc)
            
            doc.close()
            
            logger.debug(f"  PDF chunked: {len(chunks)} chunks from {file_path.name}")
            return chunks
        
        except Exception as e:
            logger.error(f"  Erreur chunking PDF {file_path.name}: {e}")
            return []
    
    def _chunk_pdf_with_toc(self, doc, toc: List) -> List[Dict]:
        """Chunke PDF selon la table des matières"""
        
        chunks = []
        
        for i, (level, title, page_num) in enumerate(toc):
            # Ignorer niveaux trop profonds
            if level > 3:
                continue
            
            # Déterminer page de fin
            if i + 1 < len(toc):
                end_page = toc[i + 1][2] - 1
            else:
                end_page = len(doc) - 1
            
            # Extraire texte de la section
            section_text = []
            for page_idx in range(page_num - 1, min(end_page, len(doc))):
                if page_idx >= 0 and page_idx < len(doc):
                    page = doc[page_idx]
                    section_text.append(page.get_text())
            
            text = ' '.join(section_text).strip()
            
            if text and len(text.split()) >= self.MIN_CHUNK_SIZE:
                chunks.append({
                    'text': text,
                    'heading': title,
                    'type': 'toc_section',
                    'pages': f"{page_num}-{end_page+1}"
                })
        
        return chunks
    
    def _chunk_pdf_structural(self, doc) -> List[Dict]:
        """Chunke PDF sur changements de structure (pages)"""
        
        chunks = []
        current_chunk = []
        
        for page_num, page in enumerate(doc):
            page_text = page.get_text()
            
            if not page_text.strip():
                continue
            
            current_chunk.append(page_text)
            
            # Créer chunk tous les 3-5 pages ou si trop gros
            chunk_text = ' '.join(current_chunk)
            word_count = len(chunk_text.split())
            
            if word_count >= self.TARGET_CHUNK_SIZE or (page_num + 1) % 4 == 0:
                if word_count >= self.MIN_CHUNK_SIZE:
                    chunks.append({
                        'text': chunk_text.strip(),
                        'heading': f"Pages {page_num - len(current_chunk) + 2}-{page_num + 1}",
                        'type': 'page_group'
                    })
                    current_chunk = []
        
        # Dernier chunk
        if current_chunk:
            chunk_text = ' '.join(current_chunk)
            if len(chunk_text.split()) >= self.MIN_CHUNK_SIZE:
                chunks.append({
                    'text': chunk_text.strip(),
                    'heading': f"Pages {len(doc) - len(current_chunk) + 1}-{len(doc)}",
                    'type': 'page_group'
                })
        
        return chunks
    
    def chunk_spreadsheet(self, file_path: Path) -> List[Dict]:
        """Chunke un tableur sémantiquement v2 : segmentation en zones + LLM Nemo"""
        
        ext = file_path.suffix.lower()
        chunks = []
        
        # Initialiser LLM si pas encore fait
        if not hasattr(self, '_llm'):
            try:
                from llm_provider import RAGConfig
                config = RAGConfig()
                self._llm = config.llm_provider
            except Exception:
                self._llm = None
        
        try:
            all_sheets_rows = []
            
            if ext == '.xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(c).strip() if c is not None else '' for c in row]
                        if any(c for c in cells):
                            rows.append(cells)
                    if rows:
                        all_sheets_rows.append((sheet_name, rows))
                wb.close()
            
            elif ext == '.ods':
                from odf import table as odf_table, text as odf_text
                from odf.opendocument import load
                
                doc = load(str(file_path))
                sheets = doc.spreadsheet.getElementsByType(odf_table.Table)
                
                for sheet in sheets:
                    sheet_name = sheet.getAttribute('name') or 'Sheet'
                    rows = []
                    for row in sheet.getElementsByType(odf_table.TableRow):
                        cell_elements = row.getElementsByType(odf_table.TableCell)
                        cells = []
                        for cell in cell_elements:
                            paras = cell.getElementsByType(odf_text.P)
                            cell_text = ' '.join([str(p) for p in paras]).strip()
                            cells.append(cell_text)
                        if any(c for c in cells):
                            rows.append(cells)
                    if rows:
                        all_sheets_rows.append((sheet_name, rows))
            
            for sheet_name, rows in all_sheets_rows:
                # Segmenter en zones
                zones = []
                current_zone = []
                for row in rows:
                    non_empty = [c for c in row if c.strip()]
                    if not non_empty:
                        if current_zone:
                            zones.append(current_zone)
                            current_zone = []
                    else:
                        current_zone.append(row)
                if current_zone:
                    zones.append(current_zone)
                
                # Fusionner zones trop petites
                merged_zones = []
                buffer = None
                for zone in zones:
                    if buffer is not None:
                        merged_zones.append(buffer + zone)
                        buffer = None
                    elif len(zone) == 1:
                        max_cell = max((len(c) for c in zone[0] if c.strip()), default=0)
                        if max_cell > 200:
                            merged_zones.append(zone)
                        else:
                            buffer = zone
                    else:
                        merged_zones.append(zone)
                if buffer:
                    if merged_zones:
                        merged_zones[-1] = merged_zones[-1] + buffer
                    else:
                        merged_zones.append(buffer)
                
                for zone_rows in merged_zones:
                    # Découper les zones trop grandes (taille dynamique ~250 mots)
                    sub_zones = []
                    total_w = sum(len(' '.join(c for c in row if c.strip()).split()) for row in zone_rows)
                    avg_wpr = max(total_w / max(len(zone_rows), 1), 1)
                    dyn_max = max(5, min(20, int(250 / avg_wpr)))
                    if len(zone_rows) <= dyn_max:
                        sub_zones = [zone_rows]
                    else:
                        first_row = zone_rows[0]
                        ne = [c for c in first_row if c.strip()]
                        is_hdr = len(ne) >= 2 and all(len(c) < 80 for c in ne)
                        nc = sum(1 for c in ne if c.replace('.','').replace(',','').replace('-','').replace('/','').strip().isdigit())
                        if nc >= len(ne) * 0.5:
                            is_hdr = False
                        hdr = zone_rows[0] if is_hdr else None
                        ds = 1 if is_hdr else 0
                        dr = zone_rows[ds:]
                        for si in range(0, len(dr), dyn_max):
                            batch = dr[si:si + dyn_max]
                            sub_zones.append(([hdr] + batch) if hdr else batch)
                    
                    for sub_zone in sub_zones:
                        # Sérialiser la zone
                        lines = [f"Feuille : « {sheet_name} »"]
                        for row in sub_zone:
                            cells = [c.strip() for c in row]
                            while cells and not cells[-1]:
                                cells.pop()
                            if cells:
                                display = [c if c else '(vide)' for c in cells]
                                lines.append(' | '.join(display))
                        
                        raw_text = '\n'.join(lines)
                        if len(raw_text.split()) < 10:
                            continue
                        
                        # Conversion LLM → dict {text, tags}
                        llm_result = self._llm_convert_table_sc(raw_text, sheet_name)
                        natural_text = llm_result['text']
                        tags = llm_result.get('tags', [])
                        
                        if not natural_text or len(natural_text.split()) < 15:
                            continue
                        
                        # Découper si trop long
                        words = natural_text.split()
                        if len(words) <= 500:
                            chunks.append({
                                'text': natural_text,
                                'heading': f"Feuille: {sheet_name}",
                                'type': 'spreadsheet',
                                'rgpd_topics': tags,
                            })
                        else:
                            paragraphs = natural_text.split('\n\n')
                            current = []
                            current_wc = 0
                            for para in paragraphs:
                                para_wc = len(para.split())
                                if current_wc + para_wc > 500 and current:
                                    chunks.append({
                                        'text': '\n\n'.join(current),
                                        'heading': f"Feuille: {sheet_name}",
                                        'type': 'spreadsheet',
                                        'rgpd_topics': tags,
                                    })
                                    current = [para]
                                    current_wc = para_wc
                                else:
                                    current.append(para)
                                    current_wc += para_wc
                            if current:
                                chunks.append({
                                    'text': '\n\n'.join(current),
                                    'heading': f"Feuille: {sheet_name}",
                                    'type': 'spreadsheet',
                                    'rgpd_topics': tags,
                                })
            
            logger.debug(f"  Spreadsheet chunked: {len(chunks)} semantic chunks from {file_path.name}")
            return chunks
        
        except Exception as e:
            logger.error(f"  Erreur chunking spreadsheet {file_path.name}: {e}")
            return []
    
    def _llm_convert_table_sc(self, raw_table_text: str, sheet_name: str) -> dict:
        """Conversion LLM pour semantic_chunker. Retourne {text, tags}."""
        PROMPT = """Tu es un assistant spécialisé en protection des données personnelles (RGPD/CNIL).

TÂCHE : Convertis ce tableau en texte naturel, lisible et complet, puis catégorise-le.

RÈGLES POUR LE TEXTE :
- Écris des phrases complètes sujet-verbe-complément
- Conserve TOUTES les informations (noms, dates, durées, chiffres, articles de loi)
- Ne résume pas, ne paraphrase pas : restitue fidèlement le contenu
- Identifie les en-têtes de colonnes ou de lignes et utilise-les comme contexte
- Si le tableau a des en-têtes en première ligne OU en première colonne, associe chaque valeur à son en-tête
- Pour les listes de valeurs, utilise des énumérations naturelles
- Ne mets pas de titre, écris directement le texte
- Écris en français (même si le tableau est en anglais, traduis)

RÈGLES POUR LES TAGS :
- À la toute fin, ajoute une ligne commençant par [TAGS] suivie des thèmes RGPD couverts
- Choisis 1 à 3 tags parmi : droits des personnes, consentement, sécurité des données, durée de conservation, sous-traitance, base légale, données sensibles, transfert hors UE, cookies, violation de données, transparence, DPO, vidéosurveillance, finalité du traitement, registre des traitements, AIPD, anonymisation, minimisation, responsable de traitement, prospection commerciale, conformité RGPD, profilage, sanctions CNIL, données de santé, information des personnes
- Si aucun ne convient, propose un tag court et descriptif

TABLEAU :
{table_text}

TEXTE NATUREL :"""
        
        if self._llm is None:
            return self._mechanical_fallback_sc(raw_table_text, sheet_name)
        
        try:
            prompt = PROMPT.format(table_text=raw_table_text)
            result = self._llm.generate(prompt, temperature=0.1, max_tokens=2000)
            result = result.strip()
            
            if 'TABLEAU :' in result or 'RÈGLES POUR' in result:
                return self._mechanical_fallback_sc(raw_table_text, sheet_name)
            
            # Détection copie brute : si le LLM a recopié le texte avec les pipes
            pipe_count = result.count(' | ')
            input_pipe_count = raw_table_text.count(' | ')
            if input_pipe_count > 3 and pipe_count > input_pipe_count * 0.5:
                logger.warning(f"LLM a recopié le tableau ({pipe_count} pipes) pour {sheet_name}, retry simplifié")
                retry_prompt = (
                    "Réécris ce contenu en français courant, sans utiliser de tableaux ni de tirets. "
                    "Chaque information doit être une phrase complète. "
                    "Conserve toutes les données (noms, dates, durées, chiffres).\n\n"
                    f"{raw_table_text}\n\nTexte réécrit :"
                )
                try:
                    result = self._llm.generate(retry_prompt, temperature=0.2, max_tokens=2000).strip()
                    pipe_count = result.count(' | ')
                    if pipe_count > input_pipe_count * 0.3:
                        return self._mechanical_fallback_sc(raw_table_text, sheet_name)
                except Exception:
                    return self._mechanical_fallback_sc(raw_table_text, sheet_name)
            
            # Extraire les tags
            tags = []
            text_lines = result.split('\n')
            clean_lines = []
            for line in text_lines:
                if line.strip().startswith('[TAGS]') or line.strip().startswith('[Tags]'):
                    tags = parse_tags(line)
                else:
                    clean_lines.append(line)
            
            clean_text = '\n'.join(clean_lines).strip()
            return {'text': f"Feuille « {sheet_name} »\n\n{clean_text}", 'tags': tags}
        except Exception:
            return self._mechanical_fallback_sc(raw_table_text, sheet_name)
    
    def _mechanical_fallback_sc(self, raw_table_text: str, sheet_name: str) -> dict:
        """Fallback mécanique si LLM indisponible : transformation propre sans pipes."""
        lines = raw_table_text.split('\n')
        text_parts = [f"Feuille « {sheet_name} »"]
        for line in lines:
            if line.startswith('Feuille :'):
                continue
            cells = [c.strip() for c in line.split(' | ')]
            cells = [c for c in cells if c and c != '(vide)']
            if not cells:
                continue
            if len(cells) == 1 and len(cells[0]) > 5:
                text_parts.append(cells[0])
            elif len(cells) == 2:
                text_parts.append(f"{cells[0]} : {cells[1]}.")
            elif len(cells) >= 3:
                values = ', '.join(cells[1:])
                text_parts.append(f"{cells[0]} : {values}.")
        return {'text': '\n'.join(text_parts), 'tags': []}
    
    def _split_by_size(self, text: str, heading: str = "") -> List[Dict]:
        """Split texte par taille cible"""
        
        words = text.split()
        chunks = []
        
        current_chunk = []
        for word in words:
            current_chunk.append(word)
            
            if len(current_chunk) >= self.TARGET_CHUNK_SIZE:
                chunks.append({
                    'text': ' '.join(current_chunk),
                    'heading': heading,
                    'type': 'size_split'
                })
                current_chunk = []
        
        # Dernier chunk
        if current_chunk and len(current_chunk) >= self.MIN_CHUNK_SIZE:
            chunks.append({
                'text': ' '.join(current_chunk),
                'heading': heading,
                'type': 'size_split'
            })
        
        return chunks
    
    def merge_small_chunks(self, chunks: List[Dict]) -> List[Dict]:
        """Fusionne les chunks trop petits"""
        
        if not chunks:
            return chunks
        
        merged = []
        buffer = None
        
        for chunk in chunks:
            word_count = len(chunk['text'].split())
            
            if word_count < self.MIN_CHUNK_SIZE:
                # Trop petit, buffer
                if buffer is None:
                    buffer = chunk
                else:
                    # Fusionner avec buffer
                    buffer['text'] = buffer['text'] + '\n\n' + chunk['text']
            else:
                # Assez grand
                if buffer:
                    # Vider buffer d'abord
                    merged.append(buffer)
                    buffer = None
                merged.append(chunk)
        
        # Dernier buffer
        if buffer:
            if merged:
                # Fusionner avec le dernier
                merged[-1]['text'] = merged[-1]['text'] + '\n\n' + buffer['text']
            else:
                merged.append(buffer)
        
        return merged
    
    def split_large_chunks(self, chunks: List[Dict]) -> List[Dict]:
        """Split les chunks trop grands"""
        
        result = []
        
        for chunk in chunks:
            word_count = len(chunk['text'].split())
            
            if word_count > self.MAX_CHUNK_SIZE:
                # Trop grand, split
                sub_chunks = self._split_by_size(chunk['text'], chunk['heading'])
                result.extend(sub_chunks)
            else:
                result.append(chunk)
        
        return result


if __name__ == "__main__":
    # Tests
    chunker = SemanticChunker()
    
    print("Chunker prêt !")
    print(f"Tailles : MIN={chunker.MIN_CHUNK_SIZE}, TARGET={chunker.TARGET_CHUNK_SIZE}, MAX={chunker.MAX_CHUNK_SIZE}")
